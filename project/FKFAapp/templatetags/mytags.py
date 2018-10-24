import datetime
# sending email
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from itertools import groupby
from os import listdir
from os.path import isfile, join

import gspread
import openpyxl
import pandas as pd
from django import template
from oauth2client.service_account import ServiceAccountCredentials


excel_path = 'C:\\Users\\Berna\\Desktop\\Fishkills.xlsx'
mypath = "C:\\Users\\Berna\\Desktop\\!\\Fish Kills\\list of fish cages"
damage_path = 'C:\\Users\\Berna\\Desktop\\fishkill damage summary .xlsx'

register = template.Library()


def OpenExcel():
    scope = ['https://www.googleapis.com/auth/spreadsheets']
    url = '/static/json/FKFAcredentials.json'
    credentials = ServiceAccountCredentials.from_json_keyfile_name(url, scope)

    gc = gspread.authorize(credentials)

    sheet = gc.open_by_key('1xHZO3FpSpJidt2VMjKo7-KuKaGJJ82SkmKLzExNjDTM')

    # print(sheet.worksheet("WQM").get_all_values())
    # print(sheet.worksheet("Sheet1").append_row(["A", "B", "asd"]))
    return sheet


@register.simple_tag
def DateList():
    print("working")
    book = openpyxl.load_workbook(excel_path, read_only=True)
    sheet = book["WQM"]
    date_list = []

    for ind, s in enumerate(sheet.rows):
        if not s[0].value: continue
        if ind < 2: continue
        date = [s[0].value, s[52].value]
        if date not in date_list:
            date_list.append(date)
    datelist = {}
    # return [datetime.datetime.strptime(d[0], "%Y-%m-%d %H:%M:%S") for d in date_list]
    for d in date_list: datelist[str(d[0].year)] = []
    for d in date_list: datelist[str(d[0].year)].append([datetime.datetime.strftime(d[0], "%Y-%m-%d"), d[1] == "FKE"])
    return datelist


@register.filter
def dateparse(date):
    return datetime.datetime.strftime(datetime.datetime.strptime(date, "%Y-%m-%d"), "%B %d, %Y")


@register.simple_tag
def get_parameter():
    book = openpyxl.load_workbook(excel_path, read_only=True)
    sheet = book["WQM"]
    return [i.value for i in list(sheet.rows)[0] if i.value]


@register.simple_tag
def param_info():
    book = openpyxl.load_workbook(excel_path, read_only=True)
    sheet = book["Parameters"]
    return [[pi.value for pi in i] for i in list(sheet.rows)[1:]]


@register.simple_tag
def acceptable(param, num, paraminfo):
    try:
        num = str(num)
        val = float(num[num.index(":") + 2:] if ":" in num else num)

        print(param, val)
        for p in paraminfo:
            if p[0].startswith(param):
                if p[2]:
                    if p[2].startswith("<"):
                        return val < float(p[2][1:])
                    elif p[2].startswith(">"):
                        return val > float(p[2][1:])
                    else:
                        if "-" in p[2]:
                            nu = p[2].split('-')
                            return float(nu[0]) <= val <= float(nu[1])
                        else:
                            return val == float(p[2])
                else:
                    return None
                break
    except (TypeError, ValueError) as asss:
        return "error"


@register.simple_tag
def get_newparam():
    return [i for i in get_parameter() if not i in ('Area', 'Date', 'Class')]


@register.simple_tag
def get_areas():
    book = openpyxl.load_workbook(excel_path, read_only=True)
    sheet = book["Areas"]
    are = []
    # index 4 = list of municipal
    for ar in list(sheet.rows)[1:]:
        are.append([ar[0].value, ar[1].value])
    are = sorted(are, key=lambda x: x[0])
    are = groupby(are, lambda x: x[0])

    ar = []
    for key, group in are:
        ar.append([key, [f[1] for f in list(group)]])  # list(group) = [Aya,['AYA','A'],[AYA,B]
    # ar.append(ar.pop([a[0] for a in ar ].index("Other")))
    return ar


@register.simple_tag
def get_headers():
    book = openpyxl.load_workbook(excel_path, read_only=True)
    sheet = book["WQM"]
    colname = ""
    param = {}
    head1 = list(sheet.rows)[0]
    head2 = list(sheet.rows)[1]
    for i, hed in enumerate(head1):
        if hed.value:
            colname = hed.value
            param[colname] = []
        param[colname].append(head2[i].value)
    return param


@register.filter('get_sub')
def get_sub(param, sub):
    return sub[param]


@register.simple_tag
def get_details(date):
    # lagay ng eror kapag wala yung date
    book = pd.read_excel(excel_path, header=[0, 1], sheet_name='WQM')

    hed = get_parameter()
    rec = []
    are = []
    for h in hed[1:]:
        a = book.loc[date][h]
        b = list(a)
        dr = []
        for c in a.itertuples():
            dr.append([r if b[i] == 'AVE' or len(b) == 1 else "{}: {}".format(b[i], r) for i, r in enumerate(c[1:]) if
                       not pd.isnull(r)])
        ad = False
        for i in dr: ad = i or ad
        if h == 'Class': continue
        if h == 'Area':
            are = [d[0] for d in dr]
        elif ad:
            rec.append([h, dr])

    return {'rec': rec, 'are': are}


# --------------------------------------------
#               CAGES
# --------------------------------------------
@register.simple_tag
def files():
    onlyfiles = [f.replace(".xlsx", "") for f in listdir(mypath) if isfile(join(mypath, f))]
    return onlyfiles


@register.filter('sheetname')
def sheetname(value):
    wb = openpyxl.load_workbook(mypath + "\\" + value + ".xlsx")
    return wb.sheetnames


@register.simple_tag
def sheetdata(muni, area):
    wb = openpyxl.load_workbook(mypath + "\\" + muni + ".xlsx")
    ws = wb[area]
    return [[ws[r][c].value for c in range(ws.max_column)] for r in range(6, ws.max_row + 1)]


# --------------------------------------------
#               / CAGES
# --------------------------------------------
@register.simple_tag
def get_damsummary():
    wb = openpyxl.load_workbook(damage_path, read_only=True)
    ws = wb.active
    return [
        [datetime.datetime.strftime(c.value, "%B %d, %Y") if isinstance(c.value, datetime.datetime) else c.value for c
         in r] for r in ws.rows]


@register.simple_tag
def dash():
    book = pd.read_excel(excel_path, header=[0, 1], sheet_name='WQM')['Dissolved Oxygen']
    h = list(book.index)
    da = []
    for ro in book.itertuples():
        if pd.isnull(ro[-1]):
            s = 0
            for i in ro[1:-1]:
                if isinstance(i, float) and not pd.isnull(i):
                    s += i
            s /= len(ro[1:-1])
        else:
            if isinstance(ro[-1], float) or isinstance(ro[-1], int):
                s = ro[-1]
            else:
                s = None
        da.append(s)

    dic = []
    for i, h1 in enumerate(h):
        if h1 not in [fi[0] for fi in dic]:
            dic.append([h1, [da[i]]])
        else:
            dic[[fi[0] for fi in dic].index(h1)][1].append(da[i])

        lbl = [datetime.datetime.strftime(ts[0], '%Y, %m, %d') for ts in dic]
    mn = [min([x for x in mnx[1] if x != None]) for mnx in dic]
    mx = [max([x for x in mnx[1] if x != None]) for mnx in dic]

    return {'mind': [[lbl[i], mn[i]] for i in range(len(lbl))], 'maxd': [[lbl[i], mx[i]] for i in range(len(lbl))]}


@register.simple_tag
def send_email(message, files):
    email_user = 'fkfabfar@gmail.com'
    email_password = 'fkfapassword'
    # email_send = '<bayen0947@gmail.com>,<tolentinojeyzellrose@gmail.com>'
    email_send = '<bayen0947@gmail.com>'

    status = []

    subject = 'subject'

    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = email_send
    msg['Subject'] = subject

    body = message
    msg.attach(MIMEText(body, 'plain'))

    # filename = ['C:\\Users\\Berna\\Downloads\\BFARlogo.png','C:\\Users\\Berna\\Downloads\\179257.jpg']
    # attachment = open(filename, 'rb')

    # part = MIMEBase('application', 'octet-stream')
    # part.set_payload((attachment).read())
    for f in files:
        try:
            # attachment = open(f, 'rb')

            part = MIMEBase('application', 'octet-stream')
            part.set_payload((f).read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', "attachment; filename= " + f.name)
            msg.attach(part)
        except:
            status.append(f.name + " failed to attach")

    # encoders.encode_base64(part)
    # part.add_header('Content-Disposition', "attachment; filename= " + filename)

    # msg.attach(part)
    text = msg.as_string()
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(email_user, email_password)

    server.sendmail(email_user, email_send, text)
    server.quit()

    return status


@register.simple_tag
def TopDamageCost():
    book = openpyxl.load_workbook(damage_path, read_only=True)
    sheet = book.active
    val = [[r[2].value, r[6].value] for i, r in enumerate(sheet.rows) if i > 0]
    val = sorted(val, key=lambda x: x[0])
    val = groupby(val, lambda x: x[0])
    val = [[r[0], sum([i[1] for i in list(r[1])])] for r in val]
    val = sorted(val, key=lambda x: x[1])
    return val[-1:-6:-1]

@register.simple_tag
def LineDamageCost():
    book = openpyxl.load_workbook(damage_path, read_only=True)
    sheet = book.active
    val = [[r[0].value, r[6].value] for i, r in enumerate(sheet.rows) if i > 0]
    val = sorted(val, key=lambda x: x[0])
    val = groupby(val, lambda x: x[0])
    val = [[datetime.datetime.strftime(r[0], '%Y, %m, %d'), sum([i[1] for i in list(r[1])])] for r in val]
    return val


@register.filter
def indexr(l, i):
    try:
        return l[i]
    except:
        return None


@register.simple_tag
def get_contact():
    book = openpyxl.load_workbook(excel_path, read_only=True)
    sheet = book["emails"]

    cont = [[ar[0].value if ar[0].value else ar[1].value, ar[1].value] for i, ar in enumerate(sheet.rows) if i > 0]
    return cont


def get_paramindex():
    book = openpyxl.load_workbook(excel_path, read_only=True)
    sheet = book["WQM"]
    return [[[pa.value, i] for i, pa in enumerate(list(sheet.rows)[0]) if pa.value], len(list(sheet.rows)[0])]


@register.simple_tag
def toarray(parname, value):
    parin = get_paramindex()
    pain = []
    for pn in parname:
        for pi in parin[0]:
            if pi[0] == pn:
                pain.append(pi)
    qw = []
    for r in value:
        dat = []
        for i in range(parin[1]): dat.append("")
        for i, vr in enumerate(r):
            for ii, val in enumerate(vr):
                dat[pain[i][1] + ii] = val
        qw.append(dat)

    # append
    book = openpyxl.load_workbook(excel_path)
    sheet = book["WQM"]

    num_rows = sheet.max_row
    for ir, ro in enumerate(qw):
        for ic, ce in enumerate(ro):
            sheet.cell(row=ir + num_rows + 1, column=ic + 1, value=ce)

    book.save(excel_path)
    return qw


@register.simple_tag
def addcontact(name, email):
    book = OpenExcel()
    sheet = book.worksheet("emails")
    sheet.append_row([name, email])



@register.simple_tag
def get_municipal():
    book = openpyxl.load_workbook(excel_path, read_only=True)
    sheet = book["Areas"]

    listbara = []
    [listbara.append(ar[0].value) for i, ar in enumerate(sheet.rows) if
     i > 0 and ar[0].value and not ar[0].value in listbara]
    print(listbara)
    return listbara
